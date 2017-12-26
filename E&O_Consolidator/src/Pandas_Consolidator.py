'''
Created on May 12, 2017

@author: v52k
'''
from os import listdir
from os.path import isfile, join, normpath
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime 
mypath = "C:/Users/v52k/OneDrive/Work/Suzhou_MM/BI/EnO_Excels/Ex 12.18.2017"
file_list = [join(normpath(mypath),f) for f in listdir(mypath) if isfile(join(mypath,f))] #Get File List

#Create Destination File

if isfile (mypath + "/Destination_Book.xlsx"):
    dest_book = openpyxl.load_workbook(filename = mypath + "/Destination_Book.xlsx")
    dest_sheet = dest_book.get_sheet_by_name("Dest_Sheet")
    attr_sheet = dest_book.get_sheet_by_name("Attr_Sheet")
    attr_df =pd.ExcelFile(mypath + "/Destination_Book.xlsx").parse("Attr_Sheet", names =['File_Name','Sheet','Product','CM','Lines','EnO'])
else:
    dest_book = openpyxl.Workbook()
    dest_sheet = dest_book.create_sheet("Dest_Sheet")
    attr_sheet = dest_book.create_sheet("Attr_Sheet")
    dest_book.save(filename = mypath + "/Destination_Book.xlsx")
    attr_df = pd.DataFrame(columns = ['File_Name','Sheet','Product','CM','Lines','EnO','TimeStamp'])

# format_df=pd.ExcelFile(mypath + "/External EnO Lancelot 05.17.2017.xlsx").parse("Load.Lancelot")

#File Loop/Copier
for f in file_list:
#     if f not in attr_df.File_Name.values:

        source_book = pd.ExcelFile(f)
        source_sheet_list = [sheet for sheet in source_book.sheet_names if ('Load' in sheet) or ('load' in sheet)]
        for sheet in source_sheet_list:
            source_df = source_book.parse(sheet)
  
    #Shape DataFrame

          
    #Get Attributes of the DataFrame
            attr_df = pd.DataFrame({'File_Name':f,
                        'Sheet':sheet,
                        'Product':source_df['Product'].iloc[0],
                        'CM':source_df['CM'].iloc[0],
                        'Lines':source_df['CM'].shape[0],
                        'EnO':source_df.sum()['Project EnO Amount'],
                        'TimeStamp':datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }, index=[0])
             
#             attr_df = [['File_Name','Sheet','Product','CM','Lines','EnO']]
          
    #Save the DataFrames in the Destination Excel 
            
    
            for r in dataframe_to_rows(source_df, index=False, header = False):
                dest_sheet.append(r)
         
            for r in dataframe_to_rows(attr_df, index=False, header=True):
                attr_sheet.append(r)
            
            dest_book.save(filename = mypath + "/Destination_Book.xlsx")         